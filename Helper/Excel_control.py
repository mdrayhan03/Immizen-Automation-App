import openpyxl
import os

class Excel_Book :
    def __init__(self, current_path) :
        self.current_dir = os.path.join('DocFile', 'Sheet', 'ExcelSheet.xlsx')

        self.exist()
    
    def exist(self) :        
        if not os.path.exists(self.current_dir) :
            self.wb = openpyxl.Workbook()
            self.sheet = self.wb.active
            self.sheet.cell(row=1, column=1, value="Client File Name")
            self.sheet.cell(row=1, column=2, value="Name")
            self.sheet.cell(row=1, column=3, value="Family Name")
            self.sheet.cell(row=1, column=4, value="Nationality")
            self.sheet.cell(row=1, column=5, value="DOB")
            self.sheet.cell(row=1, column=6, value="Phone No")
            self.sheet.cell(row=1, column=7, value="Email")
            self.sheet.cell(row=1, column=8, value="Passport No")
            self.sheet.cell(row=1, column=9, value="Address")
            self.sheet.cell(row=1, column=10, value="Application Type")
            self.sheet.cell(row=1, column=11, value="Professional Fee's")
            self.sheet.cell(row=1, column=12, value="Adminstrative Fee's")
            self.wb.save(self.current_dir)
        else :
            self.wb = openpyxl.load_workbook(self.current_dir)
            self.sheet = self.wb.active
    
    def add_info(self, arr) :
        last_row = self.sheet.max_row        
        i = 1
        for ele in arr :
            self.sheet.cell(row=last_row+1, column=i, value=ele)
            i += 1
    
    def select_info(self) :
        last_row = self.sheet.max_row
        file_name = list()
        # name = list()
        for i in range(2 , last_row+1) :
            # print(self.sheet.cell(row=i, column=1).value)
            file_name.append(self.sheet.cell(row=i, column=1).value)
            # name.append(self.sheet.calculate_dimension(row=i, column=2).value)
        
        return file_name
    
    def update_info(self, val, inp) :
        last_row = self.sheet.max_row
        row = 0
        for i in range(2, last_row+1) :
            if self.sheet.cell(row=i, column=1).value == val :
                row = i
                break
            
        j = 17
        while j == j :
            if self.sheet.cell(row=row, column=j).value == None :
                self.sheet.cell(row=row, column=j).value = inp
                break
            else :
                j += 1
                
    def save(self) :
        self.wb.save(self.current_dir)


class Milestone_Book: 
    def __init__(self, current_path) :
        self.current_dir = os.path.join('DocFile', 'Sheet', 'MilestoneSheet.xlsx')

        self.exist() 
    
    def exist(self) :
        if not os.path.exists(self.current_dir) :
            self.wb = openpyxl.Workbook()
            self.sheet = self.wb.active
            self.sheet.cell(row=1, column=1, value="Type")
            self.sheet.cell(row=1, column=2, value="Responsibility")
            self.sheet.cell(row=1, column=3, value="Milestone")
            self.save()

        else :
            self.wb = openpyxl.load_workbook(self.current_dir)
            self.sheet = self.wb.active
    
    def insert(self, arr) :
        lastrow = self.sheet.max_row

        self.sheet.cell(row=lastrow + 1, column=1, value=arr[0])
        self.sheet.cell(row=lastrow + 1, column=2, value=arr[1])

        for i in range(len(arr[2])) :
            self.sheet.cell(row=lastrow + 1, column=i+3, value=arr[2][i])

        self.save()

    def update(self, arr) :
        lastrow = self.sheet.max_row
        row = 0
        for i in range(1, lastrow + 1) :
            if self.sheet.cell(row=i, column=1).value == arr[0] :
                row = i
                break

        self.sheet.cell(row=row, column=1, value=arr[0])
        self.sheet.cell(row=row, column=2, value=arr[1])

        for i in range(len(arr[2])) :
            self.sheet.cell(row=row, column=i+3, value=arr[2][i])

        self.save()

    def select(self) :
        lastrow = self.sheet.max_row
        mainarr = list()

        for i in range(2, lastrow+1) :
            lst = list()
            for cell in self.sheet[i] :
                if cell.value is not None :
                    lst.append(cell.value)
            mainarr.append(lst)
                    
        return mainarr

    def select_type(self) :
        pass

    def save(self) :
        self.wb.save(self.current_dir)
